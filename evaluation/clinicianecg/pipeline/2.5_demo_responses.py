#!/usr/bin/env python3
#
# This source file is part of the OpenTSLM open-source project
#
# SPDX-FileCopyrightText: 2025 Stanford University, ETH Zurich, and the project authors (see CONTRIBUTORS.md)
#
# SPDX-License-Identifier: MIT
#

"""
Demo Response Generator for ECG Review System

This script generates sample responses in the Excel workbooks to demonstrate
the analysis capabilities of the response analyzer.
"""

import openpyxl
import random
import json
from pathlib import Path
import sys

def add_sample_responses(workbook_path: Path, fill_all_samples: bool = True):
    """Add sample responses to demonstrate the analysis system"""
    
    # Exact response options from the Excel workbook dropdowns (3 options each)
    response_options = {
        'pattern_recognition': [
            'Yes',
            'Some but not all', 
            'None identified'
        ],
        'clinical_reasoning': [
            'Yes',
            'Some incorrect logic',
            'Completely incorrect logic'
        ],
        'context_integration': [
            'Yes',
            'Used some key background',
            'No did not use any relevant background'
        ]
    }
    
    try:
        print(f"  Adding responses to all samples in: {workbook_path.name}")
        wb = openpyxl.load_workbook(workbook_path)
        
        # Get list of sample sheets
        sample_sheets = [name for name in wb.sheetnames if name.startswith("Sample_")]
        
        # Fill ALL sample sheets (no random selection)
        sheets_to_complete = sample_sheets
        
        for sheet_name in sheets_to_complete:
            ws = wb[sheet_name]
            
            # Add responses to the three assessment questions (H8, H15, H22)
            response_cells = ['H8', 'H15', 'H22']
            question_types = ['pattern_recognition', 'clinical_reasoning', 'context_integration']
            
            for cell_ref, question_type in zip(response_cells, question_types):
                # Select a random response with realistic distribution
                # Bias toward positive responses (reflecting typical AI model performance)
                weights = [0.5, 0.35, 0.15]  # More realistic for AI model assessment
                response = random.choices(response_options[question_type], weights=weights)[0]
                ws[cell_ref] = response
        
        # Save the workbook
        wb.save(workbook_path)
        print(f"    Added responses to {len(sheets_to_complete)} samples")
        
    except Exception as e:
        print(f"    Error adding responses to {workbook_path.name}: {e}")

def main():
    """Main function to add demo responses to all reviewer workbooks"""
    
    # Load config
    config_path = Path(__file__).parent / "config.json"
    with open(config_path, 'r') as f:
        config = json.load(f)
    
    workbooks_dir = Path("../reviewer_workbooks")
    
    print("ECG Review Demo Response Generator")
    print("=" * 40)
    print(f"Configuration: {config['reviewer_count']} reviewers, {config['reviews_per_sample']} reviews per sample")
    print("Adding complete responses to all samples for model performance analysis...")
    
    workbook_files = list(workbooks_dir.glob("Reviewer_*.xlsx"))
    
    if not workbook_files:
        print(f"No reviewer workbooks found in {workbooks_dir}")
        print("Please run the Excel generator first.")
        return
    
    print(f"Found {len(workbook_files)} reviewer workbooks")
    
    # Set random seed for reproducible demo data
    random.seed(42)
    
    for workbook_file in workbook_files:
        add_sample_responses(workbook_file, fill_all_samples=True)
    
    print("\n" + "=" * 40)
    print("Demo response generation complete!")
    print("All reviewer workbooks now contain sample responses for analysis.")
    print("You can now run the response analyzer to see the analysis capabilities.")

if __name__ == "__main__":
    main()