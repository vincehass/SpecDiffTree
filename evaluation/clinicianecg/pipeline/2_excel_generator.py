#!/usr/bin/env python3
#
# This source file is part of the OpenTSLM open-source project
#
# SPDX-FileCopyrightText: 2025 Stanford University, ETH Zurich, and the project authors (see CONTRIBUTORS.md)
#
# SPDX-License-Identifier: MIT
#

"""
ECG Review Workbook Generator - Final Version
Creates Excel files with large ECG images and optimized review layout
"""

import os
import pickle
import pandas as pd
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing import image
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

class ECGReviewGenerator:
    def __init__(self, assignments_file: str = '../reviewer_workbooks/reviewer_assignments.pkl'):
        """Initialize with reviewer assignments"""
        with open(assignments_file, 'rb') as f:
            data = pickle.load(f)
        
        self.assignments = data['assignments']
        self.sample_assignments = data['sample_assignments']
        self.samples_data = data['samples_data']
        
        # Define professional styles
        self.header_font = Font(bold=True, size=16, color="FFFFFF")
        self.section_font = Font(bold=True, size=12, color="1F4E79")
        self.normal_font = Font(size=11)
        self.small_font = Font(size=10)
        
        self.header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        self.section_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        self.input_fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")
        
        self.thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        self.medium_border = Border(
            left=Side(style='medium'), right=Side(style='medium'),
            top=Side(style='medium'), bottom=Side(style='medium')
        )
        
        # Assessment questions
        self.assessment_questions = [
            {
                'title': '1. ECG Pattern Recognition Accuracy',
                'question': 'Did the model correctly identify the relevant ECG features needed to answer the question?',
                'options': 'Yes,Some but not all,None identified'
            },
            {
                'title': '2. Clinical Reasoning Quality', 
                'question': 'Did the model appropriately connect the identified ECG features to the final answer?',
                'options': 'Yes,Some incorrect logic,Completely incorrect logic'
            },
            {
                'title': '3. Clinical Context Integration',
                'question': 'Did the model appropriately incorporate patient clinical background (age, recording conditions, artifacts) in its interpretation?',
                'options': 'Yes,Used some key background,No did not use any relevant background'
            }
        ]
    
    def get_ecg_dimensions(self, image_path: Path):
        """Get the original dimensions of the ECG image"""
        try:
            with PILImage.open(image_path) as img:
                width, height = img.size
                aspect_ratio = width / height
                return width, height, aspect_ratio
        except Exception as e:
            print(f"Error reading image dimensions: {e}")
            return 6000, 7000, 0.857  # Default ECG dimensions
    
    def crop_ecg_whitespace(self, image_path: Path, temp_dir: Path, sample_id: str = None):
        """Crop whitespace from ECG image: top 8%, bottom 7%, left 5%, right 35%"""
        try:
            with PILImage.open(image_path) as img:
                width, height = img.size
                
                # Calculate crop boundaries - enhanced cropping
                left_crop = int(width * 0.05)     # Remove left 5%
                right_crop = int(width * 0.65)    # Keep only 65% (remove right 35%)
                top_crop = int(height * 0.08)     # Remove top 8% (was 5%)
                bottom_crop = int(height * 0.93)  # Keep 93% (remove bottom 7%, was 5%)
                
                # Crop the image (left, top, right, bottom)
                cropped_img = img.crop((left_crop, top_crop, right_crop, bottom_crop))
                
                # Create unique temp filename using sample ID to avoid conflicts
                if sample_id:
                    temp_filename = f"cropped_{sample_id}_{image_path.stem}.png"
                else:
                    temp_filename = f"cropped_{image_path.parent.name}_{image_path.stem}.png"
                temp_path = temp_dir / temp_filename
                cropped_img.save(temp_path, "PNG", optimize=False)  # Keep full quality
                
                return temp_path, cropped_img.size
        except Exception as e:
            print(f"Error cropping image: {e}")
            return None, None
    
    def create_sample_sheet(self, wb: Workbook, sample_data: dict, sheet_name: str):
        """Create sample review sheet with large ECG images"""
        ws = wb.create_sheet(title=sheet_name)
        
        # Set column widths for 3-column layout with optimized ECG spacing
        ws.column_dimensions['A'].width = 35  # Clinical information
        ws.column_dimensions['B'].width = 40  # Content continuation
        ws.column_dimensions['C'].width = 3   # Spacer
        ws.column_dimensions['D'].width = 35  # ECG column (reduced for tighter spacing)
        ws.column_dimensions['E'].width = 35  # ECG continuation (reduced)
        ws.column_dimensions['F'].width = 2   # Smaller spacer between ECG and reviewer
        ws.column_dimensions['G'].width = 28  # Review questions
        ws.column_dimensions['H'].width = 25  # Review responses
        ws.column_dimensions['I'].width = 35  # Review comments
        
        current_row = 1
        
        # Header section
        ws.merge_cells(f'A{current_row}:I{current_row}')
        header_cell = ws[f'A{current_row}']
        header_cell.value = f"ECG REVIEW FORM - {sample_data['sample_id']}"
        header_cell.font = self.header_font
        header_cell.fill = self.header_fill
        header_cell.alignment = Alignment(horizontal='center', vertical='center')
        header_cell.border = self.medium_border
        ws.row_dimensions[current_row].height = 30
        current_row += 2
        
        # SECTION 1: Clinical Information (Left - Columns A-B)
        ws.merge_cells(f'A{current_row}:B{current_row}')
        info_header = ws[f'A{current_row}']
        info_header.value = "1. CLINICAL INFORMATION"
        info_header.font = self.section_font
        info_header.fill = self.section_fill
        info_header.border = self.thin_border
        current_row += 1
        
        # Sample basic info
        info_data = [
            ("Template ID:", sample_data['template_id']),
            ("ECG ID:", sample_data['ecg_id']),
            ("Sample ID:", sample_data['sample_id']),
        ]
        
        for label, value in info_data:
            ws[f'A{current_row}'] = label
            ws[f'B{current_row}'] = value
            ws[f'A{current_row}'].font = Font(bold=True, size=10)
            ws[f'A{current_row}'].border = self.thin_border
            ws[f'B{current_row}'].border = self.thin_border
            current_row += 1
        
        # Clinical Context
        clinical_start_row = current_row + 1
        ws.merge_cells(f'A{clinical_start_row}:B{clinical_start_row}')
        clinical_header = ws[f'A{clinical_start_row}']
        clinical_header.value = "Clinical Context"
        clinical_header.font = Font(bold=True, size=11, color="1F4E79")
        clinical_header.fill = self.section_fill
        clinical_header.border = self.thin_border
        current_row = clinical_start_row + 1
        
        ws.merge_cells(f'A{current_row}:B{current_row + 3}')
        clinical_cell = ws[f'A{current_row}']
        clinical_cell.value = sample_data.get('clinical_context', 'N/A')
        clinical_cell.alignment = Alignment(wrap_text=True, vertical='top')
        clinical_cell.border = self.thin_border
        clinical_cell.font = self.small_font
        for i in range(4):
            ws.row_dimensions[current_row + i].height = 18
        current_row += 4
        
        # Diagnostic Question
        question_start_row = current_row + 1
        ws.merge_cells(f'A{question_start_row}:B{question_start_row}')
        question_header = ws[f'A{question_start_row}']
        question_header.value = "Diagnostic Question"
        question_header.font = Font(bold=True, size=11, color="1F4E79")
        question_header.fill = self.section_fill
        question_header.border = self.thin_border
        current_row = question_start_row + 1
        
        ws.merge_cells(f'A{current_row}:B{current_row + 2}')
        question_cell = ws[f'A{current_row}']
        question_cell.value = sample_data.get('question', 'N/A')
        question_cell.alignment = Alignment(wrap_text=True, vertical='top')
        question_cell.border = self.thin_border
        question_cell.font = self.normal_font
        for i in range(3):
            ws.row_dimensions[current_row + i].height = 18
        current_row += 3
        
        # Answer Options (if available)
        if sample_data.get('answer_options'):
            options_start_row = current_row + 1
            ws.merge_cells(f'A{options_start_row}:B{options_start_row}')
            options_header = ws[f'A{options_start_row}']
            options_header.value = "Answer Options"
            options_header.font = Font(bold=True, size=11, color="1F4E79")
            options_header.fill = self.section_fill
            options_header.border = self.thin_border
            current_row = options_start_row + 1
            
            ws.merge_cells(f'A{current_row}:B{current_row + 3}')
            options_cell = ws[f'A{current_row}']
            options_cell.value = sample_data.get('answer_options', '')
            options_cell.alignment = Alignment(wrap_text=True, vertical='top')
            options_cell.border = self.thin_border
            options_cell.font = self.small_font
            for i in range(4):
                ws.row_dimensions[current_row + i].height = 18
            current_row += 4
        
        # Model Output
        model_start_row = current_row + 1
        ws.merge_cells(f'A{model_start_row}:B{model_start_row}')
        model_header = ws[f'A{model_start_row}']
        model_header.value = "Model Output"
        model_header.font = Font(bold=True, size=11, color="1F4E79")
        model_header.fill = self.section_fill
        model_header.border = self.thin_border
        current_row = model_start_row + 1
        
        model_output_rows = 10
        ws.merge_cells(f'A{current_row}:B{current_row + model_output_rows - 1}')
        model_cell = ws[f'A{current_row}']
        model_cell.value = sample_data.get('model_output', 'N/A')
        model_cell.alignment = Alignment(wrap_text=True, vertical='top')
        model_cell.border = self.thin_border
        model_cell.font = self.small_font
        for i in range(model_output_rows):
            ws.row_dimensions[current_row + i].height = 18
        current_row += model_output_rows
        
        # Expected Answer
        expected_start_row = current_row + 1
        ws.merge_cells(f'A{expected_start_row}:B{expected_start_row}')
        expected_header = ws[f'A{expected_start_row}']
        expected_header.value = "Expected Answer"
        expected_header.font = Font(bold=True, size=11, color="1F4E79")
        expected_header.fill = self.section_fill
        expected_header.border = self.thin_border
        current_row = expected_start_row + 1
        
        ws.merge_cells(f'A{current_row}:B{current_row}')
        expected_cell = ws[f'A{current_row}']
        expected_cell.value = sample_data.get('expected_answer', 'N/A')
        expected_cell.border = self.thin_border
        expected_cell.font = Font(bold=True, size=12)
        
        # SECTION 2: Large ECG Tracing (Center - Columns D-E)
        ecg_start_row = 3
        ws.merge_cells(f'D{ecg_start_row}:E{ecg_start_row}')
        ecg_header = ws[f'D{ecg_start_row}']
        ecg_header.value = "2. ECG TRACING"
        ecg_header.font = self.section_font
        ecg_header.fill = self.section_fill
        ecg_header.border = self.thin_border
        ecg_header.alignment = Alignment(horizontal='center')
        ecg_image_row = ecg_start_row + 1
        
        # Calculate reviewer section height to match ECG perfectly
        num_questions = len(self.assessment_questions)
        question_rows_per_assessment = 6  # title + question + response + comment rows
        reviewer_info_rows = 8  # Increased for better spacing
        total_reviewer_rows = (num_questions * question_rows_per_assessment) + reviewer_info_rows
        target_reviewer_height = total_reviewer_rows * 18  # approximately 18 pixels per row
        
        # Insert MAXIMUM SIZE ECG image filling complete column area
        try:
            if sample_data['ecg_plot_path'].exists():
                # Create temp directory for this workbook
                temp_dir = Path("./.temp")
                temp_dir.mkdir(exist_ok=True)
                
                # First crop the ECG to remove whitespace with unique identifier
                unique_id = f"{sample_data['template_id']}_{sample_data['sample_id']}"
                temp_image_path, cropped_size = self.crop_ecg_whitespace(sample_data['ecg_plot_path'], temp_dir, unique_id)
                
                if temp_image_path and temp_image_path.exists():
                    # Get cropped image dimensions
                    crop_width, crop_height = cropped_size
                    crop_aspect_ratio = crop_width / crop_height
                    
                    # Create Excel image from cropped version
                    img = image.Image(str(temp_image_path))
                    
                    # Calculate optimal size for clinical analysis (900-950px height)
                    target_display_height = min(950, max(900, target_reviewer_height * 1.1))  # 900-950px for clinical viewing
                    target_display_width = int(target_display_height * crop_aspect_ratio)
                    
                    # Set image dimensions in Excel for clinical analysis
                    img.width = target_display_width
                    img.height = target_display_height
                    
                    img.anchor = f'D{ecg_image_row}'
                    ws.add_image(img)
                    
                    # Calculate rows needed for ECG display (900-950px)
                    rows_needed = max(70, int(target_display_height / 13.5))  # Rows for large image display
                    ecg_end_row = ecg_image_row + rows_needed
                    
                    print(f"ECG Image: {crop_width}x{crop_height} (cropped, aspect: {crop_aspect_ratio:.3f}) -> {img.width}x{img.height} display (900-950px)")
                    
                else:
                    # Fallback to original image if cropping fails
                    orig_width, orig_height, aspect_ratio = self.get_ecg_dimensions(sample_data['ecg_plot_path'])
                    
                    img = image.Image(str(sample_data['ecg_plot_path']))
                    target_height = 900  # Standard fallback height
                    target_width = int(target_height * aspect_ratio)
                    
                    img.width = target_width
                    img.height = target_height
                    img.anchor = f'D{ecg_image_row}'
                    ws.add_image(img)
                    
                    rows_needed = 67  # Standard rows for 900px
                    ecg_end_row = ecg_image_row + rows_needed
                    print(f"Maximum ECG (fallback): {orig_width}x{orig_height} -> {target_width}x{target_height} (maximum size)")
            else:
                ws[f'D{ecg_image_row}'] = "ECG image not found"
                ecg_end_row = ecg_image_row + 70
        except Exception as e:
            print(f"Error adding image: {e}")
            ws[f'D{ecg_image_row}'] = f"Error loading ECG image"
            ecg_end_row = ecg_image_row + 70
        
        # SECTION 3: Review Assessment (Right - Columns G-I)
        review_start_row = 3
        ws.merge_cells(f'G{review_start_row}:I{review_start_row}')
        review_header = ws[f'G{review_start_row}']
        review_header.value = "3. REVIEWER ASSESSMENT"
        review_header.font = self.section_font
        review_header.fill = self.section_fill
        review_header.border = self.thin_border
        review_header.alignment = Alignment(horizontal='center')
        current_review_row = review_start_row + 2
        
        # Assessment questions
        for i, q_data in enumerate(self.assessment_questions):
            # Question title
            ws.merge_cells(f'G{current_review_row}:I{current_review_row}')
            title_cell = ws[f'G{current_review_row}']
            title_cell.value = q_data['title']
            title_cell.font = Font(bold=True, size=11, color="1F4E79")
            title_cell.fill = self.section_fill
            title_cell.border = self.thin_border
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            current_review_row += 1
            
            # Question text
            ws.merge_cells(f'G{current_review_row}:I{current_review_row + 1}')
            question_text_cell = ws[f'G{current_review_row}']
            question_text_cell.value = q_data['question']
            question_text_cell.alignment = Alignment(wrap_text=True, vertical='top')
            question_text_cell.border = self.thin_border
            question_text_cell.font = self.normal_font
            ws.row_dimensions[current_review_row].height = 25
            ws.row_dimensions[current_review_row + 1].height = 25
            current_review_row += 2
            
            # Response section
            ws[f'G{current_review_row}'] = "Response:"
            ws[f'G{current_review_row}'].font = Font(bold=True, size=10)
            ws[f'G{current_review_row}'].border = self.thin_border
            
            # Dropdown for responses
            response_cell = f'H{current_review_row}'
            ws[response_cell] = ""
            ws[response_cell].fill = self.input_fill
            ws[response_cell].border = self.thin_border
            
            # Create data validation for dropdown
            dv = DataValidation(type="list", formula1=f'"{q_data["options"]}"', allow_blank=True)
            dv.add(response_cell)
            ws.add_data_validation(dv)
            
            ws[f'I{current_review_row}'] = "Comments:"
            ws[f'I{current_review_row}'].font = Font(bold=True, size=10)
            ws[f'I{current_review_row}'].border = self.thin_border
            current_review_row += 1
            
            # Comment field
            comment_cell = f'G{current_review_row}'
            ws.merge_cells(f'G{current_review_row}:I{current_review_row + 1}')
            ws[comment_cell] = ""
            ws[comment_cell].fill = self.input_fill
            ws[comment_cell].border = self.thin_border
            ws[comment_cell].alignment = Alignment(wrap_text=True, vertical='top')
            ws.row_dimensions[current_review_row].height = 30
            ws.row_dimensions[current_review_row + 1].height = 30
            
            current_review_row += 3
        
        # Reviewer Information section
        current_review_row += 1
        ws.merge_cells(f'G{current_review_row}:I{current_review_row}')
        reviewer_header = ws[f'G{current_review_row}']
        reviewer_header.value = "REVIEWER INFORMATION"
        reviewer_header.font = Font(bold=True, size=11, color="1F4E79")
        reviewer_header.fill = self.section_fill
        reviewer_header.border = self.thin_border
        reviewer_header.alignment = Alignment(horizontal='center', vertical='center')
        current_review_row += 1
        
        # Auto-populated reviewer initials
        ws[f'G{current_review_row}'] = "Reviewer Initials:"
        ws[f'G{current_review_row}'].font = Font(bold=True, size=10)
        ws[f'G{current_review_row}'].border = self.thin_border
        
        ws[f'H{current_review_row}'] = "='Review_Summary'!B4"
        ws[f'H{current_review_row}'].border = self.thin_border
        ws[f'H{current_review_row}'].fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
        current_review_row += 2
        
        # Overall Comments
        ws[f'G{current_review_row}'] = "Overall Comments:"
        ws[f'G{current_review_row}'].font = Font(bold=True, size=10)
        ws[f'G{current_review_row}'].border = self.thin_border
        current_review_row += 1
        
        ws.merge_cells(f'G{current_review_row}:I{current_review_row + 2}')
        overall_comment_cell = ws[f'G{current_review_row}']
        overall_comment_cell.value = ""
        overall_comment_cell.fill = self.input_fill
        overall_comment_cell.border = self.thin_border
        overall_comment_cell.alignment = Alignment(wrap_text=True, vertical='top')
        for i in range(3):
            ws.row_dimensions[current_review_row + i].height = 25
        
        return ws
    
    def cleanup_temp_files(self):
        """Clean up temporary files and directory"""
        temp_dir = Path("./.temp")
        if temp_dir.exists():
            import shutil
            try:
                shutil.rmtree(temp_dir)
                print("Cleaned up temporary files")
            except Exception as e:
                print(f"Warning: Could not clean up temp directory: {e}")
    
    def create_summary_sheet(self, wb: Workbook, reviewer_name: str, samples: list):
        """Create summary sheet with reviewer instructions"""
        ws = wb.create_sheet(title="Review_Summary", index=0)
        
        # Set column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 35
        ws.column_dimensions['F'].width = 15
        
        # Header
        ws.merge_cells('A1:F1')
        header_cell = ws['A1']
        header_cell.value = f"ECG Review Summary - {reviewer_name}"
        header_cell.font = self.header_font
        header_cell.fill = self.header_fill
        header_cell.alignment = Alignment(horizontal='center', vertical='center')
        header_cell.border = self.medium_border
        ws.row_dimensions[1].height = 35
        
        current_row = 3
        
        # Reviewer Information Section
        ws.merge_cells(f'A{current_row}:B{current_row}')
        reviewer_info_header = ws[f'A{current_row}']
        reviewer_info_header.value = "REVIEWER INFORMATION"
        reviewer_info_header.font = self.section_font
        reviewer_info_header.fill = self.section_fill
        reviewer_info_header.border = self.thin_border
        current_row += 1
        
        ws[f'A{current_row}'] = "Enter Your Initials:"
        ws[f'A{current_row}'].font = Font(bold=True, size=12)
        ws[f'A{current_row}'].border = self.thin_border
        
        ws[f'B{current_row}'] = ""  # Cell B4 for initials
        ws[f'B{current_row}'].fill = self.input_fill
        ws[f'B{current_row}'].border = self.medium_border
        ws[f'B{current_row}'].font = Font(bold=True, size=14)
        ws[f'B{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[current_row].height = 25
        
        current_row += 1
        ws.merge_cells(f'A{current_row}:F{current_row}')
        note_cell = ws[f'A{current_row}']
        note_cell.value = "NOTE: Your initials will automatically appear on all review sheets."
        note_cell.font = Font(italic=True, size=10, color="666666")
        note_cell.alignment = Alignment(horizontal='center')
        
        current_row += 3
        
        # Instructions
        ws.merge_cells(f'A{current_row}:F{current_row}')
        instructions_header = ws[f'A{current_row}']
        instructions_header.value = "REVIEW WORKFLOW"
        instructions_header.font = self.section_font
        instructions_header.fill = self.section_fill
        instructions_header.border = self.thin_border
        current_row += 1
        
        instructions = [
            "LAYOUT: Clinical Context (left) → Large ECG Tracing (center) → Assessment Questions (right)",
            "",
            "1. Enter your initials above - they auto-populate on all review sheets",
            "2. Each sample uses optimized 3-column layout for efficient review",
            "3. ECG images are large with preserved aspect ratios for accurate analysis",
            "4. Complete 3 assessment questions using dropdown menus",
            "5. Add comments to explain your reasoning",
            "6. Save regularly and submit when completed",
            "",
            f"Total samples to review: {len(samples)}",
            "Each sample is independently reviewed by 2 expert clinicians"
        ]
        
        for instruction in instructions:
            ws.merge_cells(f'A{current_row}:F{current_row}')
            instruction_cell = ws[f'A{current_row}']
            instruction_cell.value = instruction
            if instruction.startswith("LAYOUT"):
                instruction_cell.font = Font(bold=True, size=12, color="1F4E79")
            elif instruction.startswith(("1.", "2.", "3.", "4.", "5.", "6.")):
                instruction_cell.font = Font(bold=True, size=11)
            elif instruction.startswith("Total") or instruction.startswith("Each"):
                instruction_cell.font = Font(bold=True, size=11, color="1F4E79")
            else:
                instruction_cell.font = Font(size=11)
            current_row += 1
        
        current_row += 2
        
        # Assessment Criteria
        ws.merge_cells(f'A{current_row}:F{current_row}')
        criteria_header = ws[f'A{current_row}']
        criteria_header.value = "ASSESSMENT CRITERIA"
        criteria_header.font = self.section_font
        criteria_header.fill = self.section_fill
        criteria_header.border = self.thin_border
        current_row += 1
        
        for q_data in self.assessment_questions:
            ws.merge_cells(f'A{current_row}:F{current_row}')
            question_title = ws[f'A{current_row}']
            question_title.value = q_data['title']
            question_title.font = Font(bold=True, size=11, color="1F4E79")
            current_row += 1
            
            ws.merge_cells(f'A{current_row}:F{current_row}')
            question_text = ws[f'A{current_row}']
            question_text.value = q_data['question']
            question_text.font = Font(size=10)
            current_row += 1
            
            ws.merge_cells(f'A{current_row}:F{current_row}')
            options_text = ws[f'A{current_row}']
            options_text.value = f"Options: {q_data['options'].replace(',', ' | ')}"
            options_text.font = Font(size=10, italic=True, color="666666")
            current_row += 2
        
        # Sample list
        ws.merge_cells(f'A{current_row}:F{current_row}')
        sample_list_header = ws[f'A{current_row}']
        sample_list_header.value = "ASSIGNED SAMPLES"
        sample_list_header.font = self.section_font
        sample_list_header.fill = self.section_fill
        sample_list_header.border = self.thin_border
        current_row += 1
        
        # Headers
        headers = ["Sample ID", "Template", "ECG ID", "Question Type", "Status", "Notes"]
        for i, header in enumerate(headers, start=1):
            cell = ws.cell(row=current_row, column=i)
            cell.value = header
            cell.font = Font(bold=True, size=10)
            cell.fill = self.section_fill
            cell.border = self.thin_border
            cell.alignment = Alignment(horizontal='center')
        
        current_row += 1
        
        # Sample data with automatic status tracking
        status_row_start = current_row
        for idx, sample in enumerate(samples):
            question_preview = sample.get('question', '')[:30] + '...' if sample.get('question') else 'N/A'
            
            # Create sheet name reference for status formula (must match actual sheet name)
            sheet_name = f"Sample_{idx+1:02d}"  # Use same naming as actual sheets
            
            row_data = [
                sample['sample_id'],
                f"T{sample['template_id']:02d}",
                sample['ecg_id'],
                question_preview,
                "",  # Status will be formula
                ""
            ]
            
            for i, value in enumerate(row_data, start=1):
                cell = ws.cell(row=current_row, column=i)
                if i == 5:  # Status column - add automatic tracking formula
                    # Calculate cell references for the 3 assessment questions (corrected row numbers)
                    # Question 1 response: H8, Question 2 response: H15, Question 3 response: H22
                    q1_ref = f"{sheet_name}!H8"
                    q2_ref = f"{sheet_name}!H15" 
                    q3_ref = f"{sheet_name}!H22"
                    
                    # Formula checks if all 3 assessment questions are answered
                    formula = (f'=IF(AND({q1_ref}<>"",{q2_ref}<>"",{q3_ref}<>""),'
                             f'"Completed",'
                             f'IF(OR({q1_ref}<>"",{q2_ref}<>"",{q3_ref}<>""),'
                             f'"In Progress","Not Started"))')
                    cell.value = formula
                    cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                else:
                    cell.value = value
                
                cell.border = self.thin_border
                cell.font = Font(size=10)
            
            current_row += 1
        
        # Progress summary section (after sample list)
        current_row += 1
        ws.merge_cells(f'A{current_row}:F{current_row}')
        progress_header = ws[f'A{current_row}']
        progress_header.value = "PROGRESS SUMMARY"
        progress_header.font = self.section_font
        progress_header.fill = self.section_fill
        progress_header.border = self.thin_border
        current_row += 1
        
        # Progress metrics with formulas
        start_data_row = status_row_start
        end_data_row = status_row_start + len(samples) - 1
        
        metrics = [
            ("Completed:", f'=COUNTIF(E{start_data_row}:E{end_data_row},"Completed")'),
            ("In Progress:", f'=COUNTIF(E{start_data_row}:E{end_data_row},"In Progress")'),
            ("Not Started:", f'=COUNTIF(E{start_data_row}:E{end_data_row},"Not Started")'),
            ("Total Samples:", len(samples)),
            ("Completion %:", f'=IF({len(samples)}>0,ROUND(COUNTIF(E{start_data_row}:E{end_data_row},"Completed")/{len(samples)}*100,1),0)&"%"')
        ]
        
        for metric_name, metric_formula in metrics:
            ws[f'A{current_row}'] = metric_name
            ws[f'A{current_row}'].font = Font(bold=True, size=11)
            ws[f'A{current_row}'].border = self.thin_border
            
            ws[f'B{current_row}'] = metric_formula
            ws[f'B{current_row}'].font = Font(bold=True, size=11, color="1F4E79")
            ws[f'B{current_row}'].border = self.thin_border
            if "%" in str(metric_formula):
                ws[f'B{current_row}'].fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
            current_row += 1
        
        return ws
    
    def generate_workbooks(self, output_dir: str = "../reviewer_workbooks"):
        """Generate final Excel workbooks with large ECG images"""
        output_path = Path(output_dir)
        output_path.mkdir(exist_ok=True)
        
        print(f"Generating ECG review workbooks with 900px images in {output_path}...")
        
        for reviewer_name, samples in self.assignments.items():
            print(f"Creating workbook for {reviewer_name} ({len(samples)} samples)...")
            
            # Create workbook
            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet
            
            # Create summary sheet
            self.create_summary_sheet(wb, reviewer_name, samples)
            
            # Create sample sheets with large ECG images
            for i, sample in enumerate(samples, start=1):
                sheet_name = f"Sample_{i:02d}"
                
                try:
                    self.create_sample_sheet(wb, sample, sheet_name)
                except Exception as e:
                    print(f"  Error creating sheet for {sample['sample_id']}: {e}")
                    continue
            
            # Save workbook
            filename = f"{reviewer_name}_ECG_Review.xlsx"
            filepath = output_path / filename
            
            try:
                wb.save(filepath)
                print(f"  ✓ Saved: {filepath}")
            except Exception as e:
                print(f"  ✗ Error saving {filepath}: {e}")
        
        # Clean up temporary files
        self.cleanup_temp_files()
        
        print(f"\nAll reviewer workbooks generated in: {output_path}")
        return output_path

def main():
    """Generate all reviewer workbooks"""
    # Load config
    config_path = Path(__file__).parent / "config.json"
    with open(config_path, 'r') as f:
        config = json.load(f)
    
    if not Path('../reviewer_workbooks/reviewer_assignments.pkl').exists():
        print("Error: ../reviewer_workbooks/reviewer_assignments.pkl not found. Run dataset_analyzer.py first.")
        return
    
    generator = ECGReviewGenerator()
    output_dir = generator.generate_workbooks()
    
    print(f"\n{'='*60}")
    print("ECG REVIEW WORKBOOK GENERATION COMPLETE")
    print(f"{'='*60}")
    print(f"Output directory: {output_dir}")
    print(f"Files generated: {config['reviewer_count']} reviewer workbooks")
    print(f"Reviews per sample: {config['reviews_per_sample']}")
    print(f"Features:")
    print(f"  • ECG images sized at 900-950px height for clinical analysis")
    print(f"  • Enhanced cropping: top 8%, bottom 7%, left 5%, right 35% removed")
    print(f"  • Automatic status tracking: updates as questions are answered")
    print(f"  • Progress summary with completion metrics")
    print(f"  • Optimized column spacing for better ECG-to-reviewer flow")
    print(f"  • 3-column layout: Clinical → ECG → Assessment")
    print(f"  • Automatic initials propagation")
    print(f"  • Professional styling and validation")
    print(f"{'='*60}")

if __name__ == "__main__":
    main()