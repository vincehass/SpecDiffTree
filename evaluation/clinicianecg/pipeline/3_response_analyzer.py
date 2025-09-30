#!/usr/bin/env python3
#
# This source file is part of the OpenTSLM open-source project
#
# SPDX-FileCopyrightText: 2025 Stanford University, ETH Zurich, and the project authors (see CONTRIBUTORS.md)
#
# SPDX-License-Identifier: MIT
#

"""
ECG Review Response Analyzer

This script parses the completed Excel review workbooks and generates:
- Model performance statistics and summaries
- Response distribution visualizations for manuscript figures
- Structured output in JSON, CSV, and text formats
"""

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
import json
from pathlib import Path
import openpyxl
from typing import Dict, List, Optional
from datetime import datetime
import warnings
warnings.filterwarnings('ignore')

class ECGResponseAnalyzer:
    """Analyzes responses from ECG review Excel workbooks"""
    
    def __init__(self, workbooks_dir: str = "../reviewer_workbooks"):
        self.workbooks_dir = Path(workbooks_dir)
        self.responses_data = []
        self.summary_stats = {}
        self.assessment_questions = [
            "ECG Pattern Recognition Accuracy",
            "Clinical Reasoning Quality", 
            "Clinical Context Integration"
        ]
        
    def parse_all_workbooks(self) -> pd.DataFrame:
        """Parse all Excel workbooks and extract responses"""
        print("Parsing ECG review workbooks...")
        
        all_responses = []
        workbook_files = list(self.workbooks_dir.glob("Reviewer_*.xlsx"))
        
        if not workbook_files:
            print(f"No reviewer workbooks found in {self.workbooks_dir}")
            return pd.DataFrame()
            
        for workbook_path in workbook_files:
            print(f"  Processing: {workbook_path.name}")
            reviewer_responses = self._parse_single_workbook(workbook_path)
            all_responses.extend(reviewer_responses)
            
        # Convert to DataFrame
        df = pd.DataFrame(all_responses)
        self.responses_data = df
        print(f"  Total responses extracted: {len(df)}")
        return df
    
    def _parse_single_workbook(self, workbook_path: Path) -> List[Dict]:
        """Parse a single Excel workbook and extract all responses"""
        responses = []
        
        try:
            # Load workbook
            wb = openpyxl.load_workbook(workbook_path, data_only=True)
            reviewer_name = workbook_path.stem.replace("_ECG_Review", "")
            
            # Get reviewer initials from summary sheet
            reviewer_initials = ""
            if "Review_Summary" in wb.sheetnames:
                summary_ws = wb["Review_Summary"]
                if summary_ws['B4'].value:
                    reviewer_initials = str(summary_ws['B4'].value).strip()
            
            # Parse each sample sheet
            sample_sheets = [name for name in wb.sheetnames if name.startswith("Sample_")]
            
            for sheet_name in sample_sheets:
                ws = wb[sheet_name]
                sample_response = self._extract_sample_responses(ws, reviewer_name, reviewer_initials, sheet_name)
                if sample_response:
                    responses.append(sample_response)
                    
        except Exception as e:
            print(f"    Error processing {workbook_path.name}: {e}")
            
        return responses
    
    def _extract_sample_responses(self, worksheet, reviewer_name: str, reviewer_initials: str, sheet_name: str) -> Optional[Dict]:
        """Extract responses from a single sample worksheet"""
        try:
            # Extract basic sample information
            sample_data = {
                'reviewer_name': reviewer_name,
                'reviewer_initials': reviewer_initials,
                'sheet_name': sheet_name,
                'sample_number': sheet_name.replace("Sample_", ""),
                'timestamp': datetime.now().isoformat()
            }
            
            # Extract clinical information
            template_id = worksheet['B4'].value  # Template ID is in B4
            ecg_id = worksheet['B5'].value       # ECG ID is in B5
            sample_id = worksheet['B6'].value    # Sample ID is in B6
            
            sample_data.update({
                'template_id': template_id,
                'ecg_id': ecg_id, 
                'sample_id': sample_id
            })
            
            # Extract assessment responses (H8, H15, H22)
            question_cells = ['H8', 'H15', 'H22']
            responses = []
            
            for i, cell_ref in enumerate(question_cells):
                response = worksheet[cell_ref].value
                question_name = self.assessment_questions[i]
                
                sample_data[f'question_{i+1}_response'] = response if response else ""
                sample_data[f'question_{i+1}_name'] = question_name
                
                if response:
                    responses.append(response)
            
            # Extract comments (if any)
            comment_cells = ['G9', 'G16', 'G23']  # Approximate comment locations
            for i, cell_ref in enumerate(comment_cells):
                try:
                    comment = worksheet[cell_ref].value
                    sample_data[f'question_{i+1}_comment'] = comment if comment else ""
                except:
                    sample_data[f'question_{i+1}_comment'] = ""
            
            # Calculate completion status
            completed_questions = sum(1 for r in responses if r and r.strip())
            sample_data['completion_status'] = (
                "Completed" if completed_questions == 3 else
                "In Progress" if completed_questions > 0 else
                "Not Started"
            )
            sample_data['completed_questions'] = completed_questions
            
            return sample_data
            
        except Exception as e:
            print(f"    Error extracting from {sheet_name}: {e}")
            return None
    
    def generate_summary_statistics(self) -> Dict:
        """Generate comprehensive model performance statistics"""
        if self.responses_data.empty:
            return {}
            
        df = self.responses_data
        
        stats = {
            'overview': {
                'total_samples': len(df),
                'unique_reviewers': df['reviewer_name'].nunique(),
                'total_responses': len(df[df['completion_status'] == 'Completed']),
                'response_rate': len(df[df['completion_status'] == 'Completed']) / len(df) * 100
            },
            'model_performance': {},
            'by_question': {},
            'by_reviewer': {}
        }
        
        # Model performance analysis by question
        for i in range(1, 4):
            question_col = f'question_{i}_response'
            question_name = self.assessment_questions[i-1]
            
            if question_col in df.columns:
                responses = df[df[question_col] != ''][question_col]
                total_responses = len(responses)
                
                if total_responses > 0:
                    # Calculate performance metrics
                    positive_responses = len(responses[responses == 'Yes'])
                    partial_responses = len(responses[responses.str.contains('Some|Used some', na=False)])
                    negative_responses = total_responses - positive_responses - partial_responses
                    
                    stats['model_performance'][question_name] = {
                        'total_assessments': total_responses,
                        'positive_evaluations': positive_responses,
                        'partial_evaluations': partial_responses,
                        'negative_evaluations': negative_responses,
                        'positive_rate': positive_responses / total_responses * 100,
                        'partial_rate': partial_responses / total_responses * 100,
                        'negative_rate': negative_responses / total_responses * 100,
                        'response_distribution': responses.value_counts().to_dict()
                    }
        
        # Statistics by question (detailed breakdown)
        for i in range(1, 4):
            question_col = f'question_{i}_response'
            question_name = self.assessment_questions[i-1]
            
            if question_col in df.columns:
                responses = df[df[question_col] != ''][question_col]
                stats['by_question'][question_name] = {
                    'total_responses': len(responses),
                    'response_distribution': responses.value_counts().to_dict(),
                    'response_rate': len(responses) / len(df) * 100
                }
        
        # Generate agreement analysis data needed for visualizations
        stats['agreement_analysis'] = self._calculate_basic_agreement_metrics(df)
        
        # By reviewer analysis (for understanding reviewer patterns, not completion tracking)
        for reviewer in df['reviewer_name'].unique():
            reviewer_df = df[df['reviewer_name'] == reviewer]
            completed_df = reviewer_df[reviewer_df['completion_status'] == 'Completed']
            
            if len(completed_df) > 0:
                reviewer_stats = {}
                for i in range(1, 4):
                    question_col = f'question_{i}_response'
                    if question_col in completed_df.columns:
                        responses = completed_df[question_col]
                        positive_count = len(responses[responses == 'Yes'])
                        reviewer_stats[f'question_{i}_positive_rate'] = positive_count / len(responses) * 100 if len(responses) > 0 else 0
                
                stats['by_reviewer'][reviewer] = {
                    'total_assessments': len(completed_df),
                    **reviewer_stats
                }
        
        self.summary_stats = stats
        return stats

    def _calculate_basic_agreement_metrics(self, df: pd.DataFrame) -> Dict:
        """Calculate basic agreement metrics needed for visualizations only"""
        from collections import defaultdict
        
        agreement_stats = {
            'samples_with_multiple_reviews': 0,
            'agreement_by_question': {},
            'overall_agreement_rate': 0,
            'disagreement_analysis': {}
        }
        
        # Find samples with multiple reviews
        sample_groups = df.groupby('sample_id')
        multi_reviewed_samples = [group for sample_id, group in sample_groups if len(group) > 1]
        agreement_stats['samples_with_multiple_reviews'] = len(multi_reviewed_samples)
        
        if multi_reviewed_samples:
            # Calculate agreement for each question
            for i in range(1, 4):
                question_col = f'question_{i}_response'
                question_name = self.assessment_questions[i-1]
                agreements = 0
                total_comparisons = 0
                
                # Disagreement tracking for visualization
                question_disagreement = {'disagreement_matrix': defaultdict(int)}
                
                for sample_group in multi_reviewed_samples:
                    responses = sample_group[question_col].dropna()
                    
                    if len(responses) >= 2:
                        resp_list = responses.tolist()
                        for j in range(len(resp_list)):
                            for k in range(j+1, len(resp_list)):
                                resp1, resp2 = resp_list[j], resp_list[k]
                                
                                if resp1 == resp2:
                                    agreements += 1
                                else:
                                    # Track disagreement patterns
                                    disagreement_key = self._normalize_disagreement_key(resp1, resp2)
                                    question_disagreement['disagreement_matrix'][disagreement_key] += 1
                                
                                total_comparisons += 1
                
                if total_comparisons > 0:
                    agreement_rate = agreements / total_comparisons * 100
                    agreement_stats['agreement_by_question'][question_name] = {
                        'agreement_rate': agreement_rate,
                        'total_comparisons': total_comparisons,
                        'agreements': agreements,
                        'disagreements': total_comparisons - agreements
                    }
                    agreement_stats['disagreement_analysis'][question_name] = question_disagreement
            
            # Calculate overall agreement rate
            if agreement_stats['agreement_by_question']:
                total_agreements = sum(q['agreements'] for q in agreement_stats['agreement_by_question'].values())
                total_comparisons = sum(q['total_comparisons'] for q in agreement_stats['agreement_by_question'].values())
                if total_comparisons > 0:
                    agreement_stats['overall_agreement_rate'] = total_agreements / total_comparisons * 100
        
        return agreement_stats

    def _shorten_response_label(self, response: str) -> str:
        """Create shorter, more readable labels for disagreement patterns"""
        label_mapping = {
            'Yes': 'Yes',
            'No did not use any relevant background': 'No Background',
            'Used some key background': 'Some Background',
            'Some but not all': 'Partial Recognition',
            'None identified': 'None Identified',
            'Some incorrect logic': 'Partial Logic',
            'Completely incorrect logic': 'Incorrect Logic'
        }
        
        return label_mapping.get(response, response[:20] + '...' if len(response) > 20 else response)
    
    def _is_complete_disagreement(self, response1: str, response2: str) -> bool:
        """Determine if two responses represent complete disagreement (opposite ends of scale)"""
        positive_responses = ['Yes']
        negative_responses = [
            'None identified', 
            'Completely incorrect logic', 
            'No did not use any relevant background'
        ]
        
        # Check if one is positive and the other is negative (complete disagreement)
        return ((response1 in positive_responses and response2 in negative_responses) or
                (response1 in negative_responses and response2 in positive_responses))
    

    
    def _normalize_disagreement_key(self, resp1: str, resp2: str) -> str:
        """Create normalized disagreement key to avoid duplicates like 'A vs B' and 'B vs A'"""
        # Sort responses alphabetically to ensure consistent ordering
        responses = sorted([resp1, resp2])
        return f"{responses[0]} ↔ {responses[1]}"
    
    def _create_colored_disagreement_label(self, question_part: str, resp1: str, resp2: str) -> tuple:
        """Create clean label for disagreement patterns"""
        # Get shortened labels
        resp1_short = self._shorten_response_label(resp1)
        resp2_short = self._shorten_response_label(resp2)
        
        # Create clean text label
        label = f'{question_part}\n{resp1_short} ↔ {resp2_short}'
        
        # Determine bar color based on disagreement severity
        if self._is_complete_disagreement(resp1, resp2):
            bar_color = '#e74c3c'  # Red for complete disagreement
        else:
            bar_color = '#f39c12'  # Orange for moderate disagreement
            
        return label, bar_color
    


    def create_visualizations(self, output_dir: str = "../analysis_results"):
        """Create focused visualizations for manuscript figures"""
        output_path = Path(output_dir)
        output_path.mkdir(exist_ok=True)
        
        if self.responses_data.empty:
            print("No data available for visualization")
            return
            
        df = self.responses_data
        
        # Set up the plotting style
        plt.style.use('default')
        sns.set_palette("husl")
        
        # 1. Model Assessment Distribution (model_assessment_distributions.png)
        self._plot_response_distributions(df, output_path)
        
        # 2. Agreement Rate per Question (part of agreement analysis)
        self._plot_agreement_analysis(df, output_path)
        
        # 3. Individual Subfigures (publication-ready)
        self._plot_individual_subfigures(df, output_path)
        
        print(f"Manuscript visualizations saved to: {output_path}")
    

    
    def _plot_agreement_analysis(self, df: pd.DataFrame, output_path: Path):
        """Create Agreement Rate per Question visualization"""
        if 'agreement_analysis' not in self.summary_stats:
            return
            
        agreement_data = self.summary_stats['agreement_analysis']
        
        if agreement_data['samples_with_multiple_reviews'] == 0:
            print("No samples with multiple reviews found for agreement analysis")
            return
        
        question_agreement = agreement_data.get('agreement_by_question', {})
        if not question_agreement:
            print("No question-level agreement data found")
            return
        
        # Create simple agreement rate figure
        fig, ax = plt.subplots(1, 1, figsize=(10, 6))
        
        questions = list(question_agreement.keys())
        short_questions = ['Pattern\nRecognition', 'Clinical\nReasoning', 'Context\nIntegration']
        agreement_rates = [question_agreement[q]['agreement_rate'] for q in questions]
        
        bars = ax.bar(range(len(questions)), agreement_rates, 
                      color=['#3498db', '#e74c3c', '#2ecc71'])
        ax.set_title('Agreement Rate per Question', fontweight='bold', fontsize=16, pad=20)
        ax.set_ylabel('Agreement Rate (%)', fontsize=14, fontweight='bold')
        ax.set_xlabel('Assessment Questions', fontsize=14, fontweight='bold')
        ax.set_xticks(range(len(questions)))
        ax.set_xticklabels(short_questions, fontsize=12)
        ax.set_ylim(0, 100)
        ax.grid(axis='y', alpha=0.3)
        
        # Add value labels
        for bar, rate in zip(bars, agreement_rates):
            ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 2,
                    f'{rate:.1f}%', ha='center', va='bottom', fontweight='bold', fontsize=12)
        
        plt.tight_layout()
        plt.savefig(output_path / 'agreement_rate_per_question.png', dpi=300, bbox_inches='tight')
        plt.close()
        
        print(f"✓ Agreement Rate per Question saved to: {output_path / 'agreement_rate_per_question.png'}")

    def _plot_response_distributions(self, df: pd.DataFrame, output_path: Path):
        """Create response distribution plots focused on model performance assessment"""
        fig, axes = plt.subplots(1, 3, figsize=(18, 6))
        fig.suptitle('Model Assessment Response Distributions', fontsize=16, fontweight='bold')
        
        completed_df = df[df['completion_status'] == 'Completed']
        
        for i in range(3):
            question_col = f'question_{i+1}_response'
            question_name = self.assessment_questions[i]
            
            if question_col in completed_df.columns:
                responses = completed_df[completed_df[question_col] != ''][question_col]
                
                if len(responses) > 0:
                    response_counts = responses.value_counts()
                    
                    # Define colors based on performance level
                    colors = []
                    for response in response_counts.index:
                        if response == 'Yes':
                            colors.append('#2ecc71')  # Green for positive
                        elif 'Some' in response or 'Used some' in response:
                            colors.append('#f39c12')  # Orange for partial
                        else:
                            colors.append('#e74c3c')   # Red for negative
                    
                    # Create bar plot
                    bars = axes[i].bar(range(len(response_counts)), response_counts.values, color=colors)
                    
                    # Customize the plot
                    axes[i].set_title(f'{question_name}\n({len(responses)} assessments)', fontweight='bold')
                    axes[i].set_xlabel('Assessment Response')
                    axes[i].set_ylabel('Number of Assessments')
                    axes[i].set_xticks(range(len(response_counts)))
                    
                    # Shorten labels for better display
                    short_labels = []
                    for label in response_counts.index:
                        if label == 'Yes':
                            short_labels.append('Yes')
                        elif 'Some but not all' in label:
                            short_labels.append('Some/Partial')
                        elif 'Some incorrect' in label:
                            short_labels.append('Some Incorrect')
                        elif 'Used some' in label:
                            short_labels.append('Some Context')
                        elif 'None identified' in label:
                            short_labels.append('None')
                        elif 'Completely incorrect' in label:
                            short_labels.append('Incorrect')
                        elif 'No did not use' in label:
                            short_labels.append('No Context')
                        else:
                            short_labels.append(label[:15] + '...' if len(label) > 15 else label)
                    
                    axes[i].set_xticklabels(short_labels, rotation=45, ha='right')
                    
                    # Add value labels on bars
                    for bar, value in zip(bars, response_counts.values):
                        axes[i].text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.1,
                                   str(value), ha='center', va='bottom', fontweight='bold')
                    
                    # Add percentage labels
                    total = sum(response_counts.values)
                    for bar, value in zip(bars, response_counts.values):
                        percentage = value / total * 100
                        axes[i].text(bar.get_x() + bar.get_width()/2, bar.get_height()/2,
                                   f'{percentage:.1f}%', ha='center', va='center', 
                                   color='white', fontweight='bold')
                else:
                    axes[i].text(0.5, 0.5, 'No assessments yet', ha='center', va='center',
                               transform=axes[i].transAxes, fontsize=14)
                    axes[i].set_title(question_name)
        
        plt.tight_layout()
        plt.savefig(output_path / 'model_assessment_distributions.png', dpi=300, bbox_inches='tight')
        plt.close()
    


    def save_analysis_report(self, output_dir: str = "../analysis_results"):
        """Save comprehensive analysis report"""
        output_path = Path(output_dir)
        output_path.mkdir(exist_ok=True)
        
        # Save detailed data as CSV
        if not self.responses_data.empty:
            self.responses_data.to_csv(output_path / 'detailed_responses.csv', index=False)
            print(f"Detailed responses saved to: {output_path / 'detailed_responses.csv'}")
        
        # Save summary statistics as JSON
        if self.summary_stats:
            with open(output_path / 'summary_statistics.json', 'w') as f:
                json.dump(self.summary_stats, f, indent=2)
            print(f"Summary statistics saved to: {output_path / 'summary_statistics.json'}")
        
        # Create text report
        self._create_text_report(output_path)
    
    def _create_text_report(self, output_path: Path):
        """Create a comprehensive model performance report"""
        report_file = output_path / 'model_performance_report.txt'
        
        with open(report_file, 'w') as f:
            f.write("ECG AI MODEL PERFORMANCE ANALYSIS REPORT\n")
            f.write("=" * 60 + "\n")
            f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
            
            if self.summary_stats:
                stats = self.summary_stats
                
                # Overview section
                f.write("ASSESSMENT OVERVIEW\n")
                f.write("-" * 30 + "\n")
                overview = stats['overview']
                f.write(f"Total Samples Assessed: {overview['total_responses']}\n")
                f.write(f"Total Samples in Dataset: {overview['total_samples']}\n")
                f.write(f"Number of Reviewers: {overview['unique_reviewers']}\n")
                f.write(f"Assessment Coverage: {overview['response_rate']:.1f}%\n\n")
                
                # Model performance section
                if 'model_performance' in stats:
                    f.write("MODEL PERFORMANCE SUMMARY\n")
                    f.write("-" * 40 + "\n")
                    for question, data in stats['model_performance'].items():
                        f.write(f"{question}:\n")
                        f.write(f"  - Total Assessments: {data['total_assessments']}\n")
                        f.write(f"  - Positive Evaluations: {data['positive_evaluations']} ({data['positive_rate']:.1f}%)\n")
                        f.write(f"  - Partial Evaluations: {data['partial_evaluations']} ({data['partial_rate']:.1f}%)\n")
                        f.write(f"  - Negative Evaluations: {data['negative_evaluations']} ({data['negative_rate']:.1f}%)\n")
                        
                        # Performance interpretation
                        positive_rate = data['positive_rate']
                        if positive_rate >= 70:
                            performance = "EXCELLENT"
                        elif positive_rate >= 50:
                            performance = "GOOD"
                        elif positive_rate >= 30:
                            performance = "MODERATE"
                        else:
                            performance = "NEEDS IMPROVEMENT"
                        f.write(f"  - Performance Level: {performance}\n\n")
                
                # Detailed response analysis
                f.write("DETAILED RESPONSE ANALYSIS\n")
                f.write("-" * 40 + "\n")
                for question, data in stats['by_question'].items():
                    f.write(f"{question}:\n")
                    f.write(f"  - Total Assessments: {data['total_responses']}\n")
                    f.write("  - Response Breakdown:\n")
                    for response, count in data['response_distribution'].items():
                        percentage = count / data['total_responses'] * 100
                        f.write(f"    - {response}: {count} ({percentage:.1f}%)\n")
                    f.write("\n")
                
                # Reviewer patterns
                if 'by_reviewer' in stats:
                    f.write("REVIEWER ASSESSMENT PATTERNS\n")
                    f.write("-" * 45 + "\n")
                    for reviewer, data in stats['by_reviewer'].items():
                        f.write(f"{reviewer}:\n")
                        f.write(f"  - Total Assessments: {data['total_assessments']}\n")
                        for key, value in data.items():
                            if 'positive_rate' in key:
                                question_num = key.split('_')[1]
                                f.write(f"  - Question {question_num} Positive Rate: {value:.1f}%\n")
                        f.write("\n")
                
                # Summary and recommendations
                f.write("SUMMARY AND RECOMMENDATIONS\n")
                f.write("-" * 45 + "\n")
                
                if 'model_performance' in stats:
                    # Calculate overall model performance
                    total_positive = sum(data['positive_evaluations'] for data in stats['model_performance'].values())
                    total_assessments = sum(data['total_assessments'] for data in stats['model_performance'].values())
                    overall_positive_rate = (total_positive / total_assessments * 100) if total_assessments > 0 else 0
                    
                    f.write(f"Overall Model Positive Assessment Rate: {overall_positive_rate:.1f}%\n\n")
                    
                    # Identify strengths and weaknesses
                    strengths = []
                    weaknesses = []
                    
                    for question, data in stats['model_performance'].items():
                        if data['positive_rate'] >= 60:
                            strengths.append(question)
                        elif data['positive_rate'] < 40:
                            weaknesses.append(question)
                    
                    if strengths:
                        f.write("Model Strengths:\n")
                        for strength in strengths:
                            f.write(f"  - {strength}\n")
                        f.write("\n")
                    
                    if weaknesses:
                        f.write("Areas for Improvement:\n")
                        for weakness in weaknesses:
                            f.write(f"  - {weakness}\n")
                        f.write("\n")
        
        print(f"Model performance report saved to: {report_file}")

    def _plot_individual_subfigures(self, df: pd.DataFrame, output_path: Path):
        """Generate each subfigure as a separate PDF for flexible manuscript use"""
        
        completed_df = df[df['completion_status'] == 'Completed']
        question_names = ['Pattern\nRecognition', 'Clinical\nReasoning', 'Context\nIntegration']
        
        # SUBFIGURE A: Performance by Assessment Area
        fig_a, ax = plt.subplots(1, 1, figsize=(7, 4.5))  # Increased width from 6 to 7
        
        positive_rates = []
        partial_rates = []
        negative_rates = []
        
        for i in range(1, 4):
            question_col = f'question_{i}_response'
            if question_col in completed_df.columns:
                responses = completed_df[question_col]
                if len(responses) > 0:
                    positive = len(responses[responses == 'Yes']) / len(responses) * 100
                    partial = len(responses[responses.str.contains('Some|Used some', na=False)]) / len(responses) * 100
                    negative = 100 - positive - partial
                    
                    positive_rates.append(positive)
                    partial_rates.append(partial)
                    negative_rates.append(negative)
                else:
                    positive_rates.append(0)
                    partial_rates.append(0)
                    negative_rates.append(0)
        
        x = np.arange(len(question_names))
        width = 0.6
        
        bars1 = ax.bar(x, positive_rates, width, label='Positive', color='#2ecc71')
        bars2 = ax.bar(x, partial_rates, width, bottom=positive_rates, label='Partial', color='#f39c12')
        bars3 = ax.bar(x, negative_rates, width, bottom=np.array(positive_rates) + np.array(partial_rates), 
                       label='Negative', color='#e74c3c')
        
        ax.set_xlabel('Assessment Area', fontweight='bold', fontsize=18)  # Increased from 16 to 18
        ax.set_ylabel('Percentage (%)', fontweight='bold', fontsize=18)  # Increased from 16 to 18
        ax.set_xticks(x)
        ax.set_xticklabels(question_names, rotation=0, ha='center', fontsize=16)  # Removed rotation, increased font
        ax.legend(loc='lower center', bbox_to_anchor=(0.5, -0.03), ncol=3, fontsize=16)  # Increased from 14 to 16
        ax.set_ylim(0, 100)
        ax.grid(axis='y', alpha=0.3)
        ax.tick_params(axis='y', labelsize=16)  # Increased from 14 to 16
        
        # Add percentage labels on bars
        for i, (bar1, bar2, bar3) in enumerate(zip(bars1, bars2, bars3)):
            if positive_rates[i] > 1:  # Show label for any segment larger than 1%
                ax.text(bar1.get_x() + bar1.get_width()/2, bar1.get_height()/2,
                        f'{positive_rates[i]:.0f}%', ha='center', va='center', 
                        fontweight='bold', color='white' if positive_rates[i] > 15 else 'black', fontsize=12)
            if partial_rates[i] > 1:
                ax.text(bar2.get_x() + bar2.get_width()/2, 
                        positive_rates[i] + bar2.get_height()/2,
                        f'{partial_rates[i]:.0f}%', ha='center', va='center', 
                        fontweight='bold', color='white' if partial_rates[i] > 15 else 'black', fontsize=12)
            if negative_rates[i] > 0.5:  # Lower threshold for negative rates to catch small values like 1.2%
                ax.text(bar3.get_x() + bar3.get_width()/2, 
                        positive_rates[i] + partial_rates[i] + bar3.get_height()/2,
                        f'{negative_rates[i]:.0f}%', ha='center', va='center', 
                        fontweight='bold', color='white' if negative_rates[i] > 15 else 'black', fontsize=12)
        
        plt.tight_layout()
        plt.savefig(output_path / 'subfigure_A_performance_by_area.pdf', dpi=300, bbox_inches='tight')
        plt.close()
        
        # SUBFIGURE B: Assessment Patterns by Reviewer
        fig_b, ax = plt.subplots(1, 1, figsize=(7, 4.5))  # Increased width from 6 to 7
        
        reviewers = sorted(completed_df['reviewer_name'].unique())  # Sort alphabetically
        reviewer_labels = [chr(65 + i) for i in range(len(reviewers))]  # A, B, C, D, E
        
        # Create assessment matrix for heatmap
        assessment_matrix = []
        for reviewer in reviewers:
            reviewer_df = completed_df[completed_df['reviewer_name'] == reviewer]
            reviewer_rates = []
            
            for i in range(1, 4):
                question_col = f'question_{i}_response'
                if question_col in reviewer_df.columns:
                    responses = reviewer_df[question_col].dropna()
                    if len(responses) > 0:
                        positive_rate = len(responses[responses == 'Yes']) / len(responses)
                        reviewer_rates.append(positive_rate)
                    else:
                        reviewer_rates.append(0)
                else:
                    reviewer_rates.append(0)
            
            assessment_matrix.append(reviewer_rates)
        
        if assessment_matrix and len(reviewers) > 0:
            im = ax.imshow(assessment_matrix, cmap='RdYlGn', aspect='auto', vmin=0, vmax=1)
            
            ax.set_xlabel('Assessment Area', fontweight='bold', fontsize=18)  # Increased from 16 to 18
            ax.set_ylabel('Reviewers', fontweight='bold', fontsize=18)  # Increased from 16 to 18
            
            # Set ticks and labels
            ax.set_xticks(range(3))
            ax.set_xticklabels(['Pattern\nRecognition', 'Clinical\nReasoning', 'Context\nIntegration'], fontsize=16)  # Increased from 14 to 16
            ax.set_yticks(range(len(reviewers)))
            ax.set_yticklabels(reviewer_labels, fontsize=16)  # Use A, B, C, D, E labels
            ax.tick_params(axis='both', labelsize=16)  # Increased from 14 to 16
            
            # Add colorbar
            cbar = plt.colorbar(im, ax=ax)
            cbar.set_label('Positive Assessment Rate', rotation=270, labelpad=20, fontsize=16)  # Increased from 14 to 16
            cbar.ax.tick_params(labelsize=14)  # Increased from 12 to 14
        else:
            ax.text(0.5, 0.5, 'No reviewer data available', ha='center', va='center', 
                    transform=ax.transAxes, fontsize=16)
        
        plt.tight_layout()
        plt.savefig(output_path / 'subfigure_B_reviewer_patterns.pdf', dpi=300, bbox_inches='tight')
        plt.close()
        
        # SUBFIGURE C: Disagreement Patterns
        fig_c, ax = plt.subplots(1, 1, figsize=(9, 7))  # Increased width from 8 to 9 and height from 6 to 7
        
        if 'agreement_analysis' in self.summary_stats:
            agreement_data = self.summary_stats['agreement_analysis']
            disagreement_data = agreement_data.get('disagreement_analysis', {})
            
            if disagreement_data:
                # Process disagreements by question
                question_disagreements = {}
                questions = ['ECG Pattern Recognition Accuracy', 'Clinical Reasoning Quality', 'Clinical Context Integration']
                
                for i, question in enumerate(questions):
                    if question in disagreement_data:
                        question_matrix = disagreement_data[question].get('disagreement_matrix', {})
                        question_short = ['Q1: Pattern', 'Q2: Reasoning', 'Q3: Context'][i]
                        
                        for pattern, count in question_matrix.items():
                            if count > 0:
                                pattern_with_question = f"{question_short} - {pattern}"
                                question_disagreements[pattern_with_question] = count
                
                if question_disagreements:
                    patterns = list(question_disagreements.keys())
                    counts = list(question_disagreements.values())
                    
                    # Sort by frequency and take top 12 for readability
                    sorted_pairs = sorted(zip(patterns, counts), key=lambda x: x[1], reverse=True)
                    top_patterns = sorted_pairs[:12]
                    patterns, counts = zip(*top_patterns) if top_patterns else ([], [])
                    
                    # Create cleaner labels
                    clean_labels = []
                    bar_colors = []
                    
                    for p in patterns:
                        if ' - ' in p and ('↔' in p or ' vs ' in p):
                            question_part, disagreement_part = p.split(' - ', 1)
                            
                            # Handle both old and new formats
                            if '↔' in disagreement_part:
                                parts = disagreement_part.split(' ↔ ')
                                response1 = parts[0].strip()
                                response2 = parts[1].strip()
                            elif ' vs ' in disagreement_part:
                                parts = disagreement_part.split(' vs ')
                                response1 = parts[0].strip()
                                response2 = parts[1].strip()
                            else:
                                clean_labels.append(p)
                                bar_colors.append('#95a5a6')
                                continue
                            
                            # Use the new colored label creation
                            label, bar_color = self._create_colored_disagreement_label(question_part, response1, response2)
                            clean_labels.append(label)
                            bar_colors.append(bar_color)
                        else:
                            clean_labels.append(p)
                            bar_colors.append('#95a5a6')
                    
                    # Create horizontal bar chart
                    y_positions = np.arange(len(patterns))
                    bars = ax.barh(y_positions, counts, color=bar_colors, height=0.7)
                    
                    ax.set_xlabel('Frequency', fontweight='bold', fontsize=18)  # Increased from 16 to 18
                    ax.set_yticks(y_positions)
                    
                    # Set labels with larger font size
                    ax.set_yticklabels(clean_labels, fontsize=16, ha='right')  # Increased from 14 to 16
                    
                    ax.tick_params(axis='x', labelsize=16)  # Increased from 14 to 16
                    
                    # Add value labels
                    for i, (bar, count) in enumerate(zip(bars, counts)):
                        ax.text(bar.get_width() + max(counts) * 0.02, bar.get_y() + bar.get_height()/2,
                                str(count), ha='left', va='center', fontweight='bold', fontsize=14)  # Increased from 12 to 14
                    
                    # Add legend with clearer labels
                    from matplotlib.patches import Rectangle
                    legend_elements = [
                        Rectangle((0, 0), 1, 1, facecolor='#e74c3c', label='Complete (Yes ↔ No)'),
                        Rectangle((0, 0), 1, 1, facecolor='#f39c12', label='Moderate (Adjacent)'),
                    ]
                    # Position legend in bottom right with custom positioning to avoid overlap
                    ax.legend(handles=legend_elements, loc='lower right', bbox_to_anchor=(1.0, 0.02), fontsize=14)  # Smaller font and custom position
                    
                    ax.set_xlim(0, max(counts) * 1.3)  # Increased to 1.3 for more space
                    ax.grid(axis='x', alpha=0.3)
                    ax.invert_yaxis()
                else:
                    ax.text(0.5, 0.5, 'No disagreement patterns found', ha='center', va='center', 
                            transform=ax.transAxes, fontsize=12)

            else:
                ax.text(0.5, 0.5, 'No disagreement analysis available', ha='center', va='center', 
                        transform=ax.transAxes, fontsize=12)

        else:
            ax.text(0.5, 0.5, 'Agreement analysis not available', ha='center', va='center', 
                    transform=ax.transAxes, fontsize=12)

        
        plt.tight_layout()
        plt.savefig(output_path / 'subfigure_C_disagreement_patterns.pdf', dpi=300, bbox_inches='tight')
        plt.close()
        
        print(f"✓ Individual subfigures saved:")
        print(f"  - Subfigure A: {output_path / 'subfigure_A_performance_by_area.pdf'}")
        print(f"  - Subfigure B: {output_path / 'subfigure_B_reviewer_patterns.pdf'}")
        print(f"  - Subfigure C: {output_path / 'subfigure_C_disagreement_patterns.pdf'}")

def main():
    """Main function to run the analysis"""
    # Load config
    config_path = Path(__file__).parent / "config.json"
    with open(config_path, 'r') as f:
        config = json.load(f)
    
    print("ECG Review Response Analyzer")
    print("=" * 40)
    print(f"Configuration: {config['reviewer_count']} reviewers, {config['reviews_per_sample']} reviews per sample")
    
    # Initialize analyzer
    analyzer = ECGResponseAnalyzer()
    
    # Parse all workbooks
    responses_df = analyzer.parse_all_workbooks()
    
    if responses_df.empty:
        print("No responses found. Make sure reviewer workbooks exist and contain responses.")
        return
    
    # Generate statistics
    print("\nGenerating summary statistics...")
    stats = analyzer.generate_summary_statistics()
    
    # Create visualizations
    print("Creating visualizations...")
    analyzer.create_visualizations()
    
    # Save comprehensive report
    print("Saving analysis report...")
    analyzer.save_analysis_report()
    
    print("\nAnalysis complete!")
    print(f"Results saved in: ../analysis_results/")
    
    # Print quick summary
    if stats:
        overview = stats['overview']
        print(f"\nModel Performance Summary:")
        print(f"- {overview['total_responses']} completed assessments")
        print(f"- {overview['unique_reviewers']} expert reviewers")
        print(f"- {overview['response_rate']:.1f}% of dataset assessed")
        
        # Add model performance highlights
        if 'model_performance' in stats:
            total_positive = sum(data['positive_evaluations'] for data in stats['model_performance'].values())
            total_assessments = sum(data['total_assessments'] for data in stats['model_performance'].values())
            if total_assessments > 0:
                overall_positive_rate = total_positive / total_assessments * 100
                print(f"- {overall_positive_rate:.1f}% overall positive assessment rate")

if __name__ == "__main__":
    main()